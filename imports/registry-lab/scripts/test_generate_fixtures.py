#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "pyarrow>=16",
# ]
# ///
"""Focused tests for decentralized demo fixture alignment."""

from __future__ import annotations

import importlib.util
import unittest
import csv
import datetime as dt
import re
import sys
from pathlib import Path

import pyarrow.parquet as pq
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
GENERATOR_PATH = SCRIPT_DIR / "generate-fixtures.py"
AGRI_GENERATOR_PATH = SCRIPT_DIR / "generate-agri-fixtures.py"
SCRIPT_MATRIX_PATHS = [
    SCRIPT_DIR / "demo-flow.py",
    SCRIPT_DIR / "demo-live-stories.py",
    SCRIPT_DIR / "smoke-notary-client.py",
]


def load_generator():
    return load_module(GENERATOR_PATH)


def load_module(path: Path):
    spec = importlib.util.spec_from_file_location(path.stem.replace("-", "_"), path)
    if not spec or not spec.loader:
        raise RuntimeError(f"could not load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class GenerateFixturesTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.generator = load_generator()
        cls.expected_outcomes = {
            "NID-1001": {"alive": True, "health": True, "combined": True},
            "NID-1002": {"alive": True, "health": False, "combined": False},
            "NID-1003": {"alive": False, "health": True, "combined": False},
            "NID-1004": {"alive": True, "health": True, "combined": True},
            "NID-1005": {"alive": True, "health": False, "combined": False},
            "NID-1006": {"alive": True, "health": True, "combined": True},
            "NID-1007": {"alive": True, "health": True, "combined": False},
            "NID-1008": {"alive": True, "health": True, "combined": True},
            "NID-1009": {"alive": True, "health": True, "combined": False},
            "NID-1010": {"alive": True, "health": False, "combined": False},
        }

    def test_v1_registry_lab_matrix_matches_openspp_story_people(self) -> None:
        expected = {
            "NID-1001": ("Miguel", "Santos", "2016-01-15", "child", "false"),
            "NID-1002": ("Maria", "Dela Cruz", "2018-01-15", "child", "false"),
            "NID-1003": ("Cara", "Okafor", "1957-02-14", "adult", "true"),
            "NID-1004": ("Rafael", "Aquino", "2019-01-15", "child", "false"),
            "NID-1005": ("Rosalie", "Bautista", "2013-01-15", "child", "false"),
            "NID-1006": ("Miguel", "Martinez", "2014-01-15", "child", "false"),
            "NID-1007": ("Lola", "Santos", "1958-01-15", "elderly", "false"),
            "NID-1008": ("Rosa", "Garcia", "1954-01-15", "elderly", "false"),
            "NID-1009": ("Ana", "Mendoza", "1998-01-15", "adult", "false"),
            "NID-1010": ("Pedro", "Reyes", "1971-01-15", "adult", "false"),
        }
        civil_by_id = {
            row[0]: {
                "given_name": row[1],
                "surname": row[2],
                "birth_date": row[3],
                "life_stage": row[4],
                "deceased": row[5],
            }
            for row in self.generator.data_rows(self.generator.CIVIL_ROWS)
        }

        self.assertEqual(set(expected), set(civil_by_id) & set(expected))
        for national_id, (given_name, surname, birth_date, life_stage, deceased) in expected.items():
            with self.subTest(national_id=national_id):
                self.assertEqual(civil_by_id[national_id]["given_name"], given_name)
                self.assertEqual(civil_by_id[national_id]["surname"], surname)
                self.assertEqual(civil_by_id[national_id]["birth_date"], birth_date)
                self.assertEqual(civil_by_id[national_id]["life_stage"], life_stage)
                self.assertEqual(civil_by_id[national_id]["deceased"], deceased)

    def test_guided_demo_personas_have_credible_source_rows(self) -> None:
        civil_by_id = {
            row[0]: {
                "given_name": row[1],
                "surname": row[2],
                "birth_date": row[3],
                "life_stage": row[4],
                "deceased": row[5],
                "district": row[6],
            }
            for row in self.generator.data_rows(self.generator.CIVIL_ROWS)
        }
        enrollment_by_id = {row[3]: row for row in self.generator.data_rows(self.generator.ENROLLMENTS)}
        health_by_id = {row["national_id"]: row for row in self.generator.HEALTH_ROWS}
        agri = load_module(AGRI_GENERATOR_PATH)
        farmers_by_id = {row["farmer_id"]: row for row in agri.FARMERS}

        self.assertEqual(
            civil_by_id["NID-1001"],
            {
                "given_name": "Miguel",
                "surname": "Santos",
                "birth_date": "2016-01-15",
                "life_stage": "child",
                "deceased": "false",
                "district": "north",
            },
        )
        self.assertEqual(civil_by_id["NID-2001"]["life_stage"], "adult")
        self.assertEqual(civil_by_id["NID-2001"]["deceased"], "false")
        self.assertEqual(enrollment_by_id["NID-1001"][4:6], ["CHILD_SUPPORT", "active"])
        self.assertTrue(self._health_available(health_by_id["NID-1001"]))
        self.assertFalse(self._health_available(health_by_id["NID-1002"]))
        self.assertEqual(farmers_by_id["FARMER-1001"]["given_name"], "Amina")
        self.assertEqual(farmers_by_id["FARMER-1001"]["family_name"], "Kone")
        self.assertEqual(farmers_by_id["FARMER-1002"]["given_name"], "Bako")
        self.assertEqual(farmers_by_id["FARMER-1003"]["given_name"], "Chipo")

    def test_public_demo_names_do_not_regress_to_misleading_civil_status(self) -> None:
        self.assertIn("life_stage", self.generator.CIVIL_ROWS[0])
        self.assertNotIn("civil_status", self.generator.CIVIL_ROWS[0])

        public_files = [
            self.generator.ROOT / "config" / "lab-homepage" / "public-demo-credentials.json",
            self.generator.ROOT / "config" / "coolify" / "relay" / "civil-registry-relay.yaml",
            self.generator.ROOT / "config" / "coolify" / "relay" / "civil-registry-relay.metadata.yaml",
            self.generator.ROOT / "config" / "static-metadata" / "metadata.yaml",
        ]
        for path in public_files:
            with self.subTest(path=path.relative_to(self.generator.ROOT)):
                text = path.read_text(encoding="utf-8")
                self.assertNotIn("civil_status", text)
                self.assertNotIn("civil-status", text)

    def test_published_metadata_omits_private_notary_hostnames(self) -> None:
        public_metadata_paths = [
            self.generator.ROOT / "config" / "static-metadata" / "metadata.yaml",
            *sorted((self.generator.ROOT / "config" / "coolify" / "relay").glob("*.metadata.yaml")),
        ]
        private_hosts = (
            "civil-notary:8080",
            "social-protection-notary:8080",
            "shared-eligibility-notary:8080",
            "nagdi-agriculture-notary:8080",
        )
        for path in public_metadata_paths:
            with self.subTest(path=path.relative_to(self.generator.ROOT)):
                text = path.read_text(encoding="utf-8")
                for host in private_hosts:
                    self.assertNotIn(host, text)

    def test_v1_notary_outcomes_are_encoded_by_fixture_facts(self) -> None:
        civil_by_id = {row[0]: row for row in self.generator.data_rows(self.generator.CIVIL_ROWS)}
        enrollment_by_id = {row[3]: row for row in self.generator.data_rows(self.generator.ENROLLMENTS)}
        health_by_id = {row["national_id"]: row for row in self.generator.HEALTH_ROWS}

        for national_id, expected in self.expected_outcomes.items():
            with self.subTest(national_id=national_id):
                alive = civil_by_id[national_id][5] == "false"
                health = self._health_available(health_by_id[national_id])
                social_active = enrollment_by_id.get(national_id, [None] * 6)[5] == "active"
                self.assertEqual(alive, expected["alive"])
                self.assertEqual(health, expected["health"])
                self.assertEqual(alive and health and social_active, expected["combined"])

    def test_fixture_relationships_remain_valid(self) -> None:
        self.generator.validate_fixture_coverage()

    def test_generated_outputs_match_fixture_source(self) -> None:
        civil_path = self.generator.DATA_DIR / "civil" / "civil-persons.csv"
        social_path = self.generator.DATA_DIR / "social-protection" / "social-protection.xlsx"
        health_path = self.generator.DATA_DIR / "health" / "health-facilities.parquet"

        with civil_path.open(newline="", encoding="utf-8") as handle:
            civil_rows = list(csv.reader(handle))
        self.assertEqual(civil_rows, self.generator.CIVIL_ROWS)

        workbook = load_workbook(social_path, data_only=True)
        expected_sheets = {
            "Households": self.generator.HOUSEHOLDS,
            "Persons": self.generator.PERSONS,
            "Enrollments": self.generator.ENROLLMENTS,
            "DistrictGeometries": self.generator.DISTRICT_GEOMETRIES,
        }
        for sheet_name, expected_rows in expected_sheets.items():
            with self.subTest(sheet_name=sheet_name):
                sheet = workbook[sheet_name]
                actual_rows = [
                    [value.date() if isinstance(value, dt.datetime) else value for value in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                self.assertEqual(actual_rows, expected_rows)

        health_table = pq.read_table(health_path)
        health_rows = health_table.to_pylist()
        self.assertEqual(health_rows, self.generator.HEALTH_ROWS)

    def test_script_v1_matrices_match_fixture_outcomes(self) -> None:
        for path in SCRIPT_MATRIX_PATHS:
            with self.subTest(path=path.name):
                module = load_module(path)
                matrix = {
                    item["id"]: {
                        "alive": item["alive"],
                        "health": item["health"],
                        "combined": item["combined"],
                    }
                    for item in module.V1_MATRIX
                }
                self.assertEqual(matrix, self.expected_outcomes)

        smoke_matrix = self._smoke_shell_matrix()
        for national_id, expected in self.expected_outcomes.items():
            with self.subTest(path="smoke.sh", national_id=national_id):
                self.assertEqual(smoke_matrix[("person-is-alive", national_id)], expected["alive"])
                self.assertEqual(smoke_matrix[("health-service-available", national_id)], expected["health"])
                self.assertEqual(smoke_matrix[("eligible-for-combined-support", national_id)], expected["combined"])

    @staticmethod
    def _health_available(row: dict[str, object]) -> bool:
        return (
            row["license_status"] == "active"
            and row["pediatric_service_available"] is True
            and row["practitioner_credential_active"] is True
        )

    @staticmethod
    def _smoke_shell_matrix() -> dict[tuple[str, str], bool]:
        text = (SCRIPT_DIR / "smoke.sh").read_text(encoding="utf-8")
        entries = re.findall(r'"[^"]+\|[^"]+\|[^"]+\|([^|"]+)\|(NID-\d+)\|(true|false)"', text)
        return {(claim, national_id): expected == "true" for claim, national_id, expected in entries}


if __name__ == "__main__":
    unittest.main()
