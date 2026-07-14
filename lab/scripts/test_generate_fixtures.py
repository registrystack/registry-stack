#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "pyarrow>=16",
#   "PyYAML>=6.0",
# ]
# ///
"""Focused tests for decentralized demo fixture alignment."""

from __future__ import annotations

import importlib.util
import unittest
import csv
import datetime as dt
import sys
import tempfile
from pathlib import Path

import pyarrow.parquet as pq
import yaml
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
GENERATOR_PATH = SCRIPT_DIR / "generate-fixtures.py"
AGRI_GENERATOR_PATH = SCRIPT_DIR / "generate-agri-fixtures.py"


def load_generator():
    return load_module(GENERATOR_PATH)


def load_module(path: Path):
    spec = importlib.util.spec_from_file_location(path.stem.replace("-", "_"), path)
    if not spec or not spec.loader:
        raise RuntimeError(f"could not load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class GenerateFixturesTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.generator = load_generator()
        cls.expected_outcomes = {
            "NID-1001": {"alive": True, "health": True, "combined": True},
            "NID-1002": {"alive": True, "health": False, "combined": False},
            "NID-1003": {"alive": False, "health": True, "combined": False},
            "NID-1004": {"alive": True, "health": True, "combined": True},
            "NID-1005": {"alive": True, "health": False, "combined": False},
            "NID-1006": {"alive": True, "health": True, "combined": True},
            "NID-1007": {"alive": True, "health": True, "combined": False},
            "NID-1008": {"alive": True, "health": True, "combined": True},
            "NID-1009": {"alive": True, "health": True, "combined": False},
            "NID-1010": {"alive": True, "health": False, "combined": False},
        }
        cls.expected_runtime_outcomes = {
            national_id: dict(expected)
            for national_id, expected in cls.expected_outcomes.items()
        }

    def test_v1_registry_lab_matrix_matches_openspp_story_people(self) -> None:
        expected = {
            "NID-1001": ("Miguel", "Santos", "2016-01-15", "child", "false"),
            "NID-1002": ("Maria", "Dela Cruz", "2018-01-15", "child", "false"),
            "NID-1003": ("Cara", "Okafor", "1957-02-14", "adult", "true"),
            "NID-1004": ("Rafael", "Aquino", "2019-01-15", "child", "false"),
            "NID-1005": ("Rosalie", "Bautista", "2013-01-15", "child", "false"),
            "NID-1006": ("Miguel", "Martinez", "2014-01-15", "child", "false"),
            "NID-1007": ("Lola", "Santos", "1958-01-15", "elderly", "false"),
            "NID-1008": ("Rosa", "Garcia", "1954-01-15", "elderly", "false"),
            "NID-1009": ("Ana", "Mendoza", "1998-01-15", "adult", "false"),
            "NID-1010": ("Pedro", "Reyes", "1971-01-15", "adult", "false"),
        }
        civil_by_id = {
            row[0]: {
                "given_name": row[1],
                "surname": row[2],
                "birth_date": row[3],
                "life_stage": row[4],
                "deceased": row[5],
            }
            for row in self.generator.data_rows(self.generator.CIVIL_ROWS)
        }

        self.assertEqual(set(expected), set(civil_by_id) & set(expected))
        for national_id, (given_name, surname, birth_date, life_stage, deceased) in expected.items():
            with self.subTest(national_id=national_id):
                self.assertEqual(civil_by_id[national_id]["given_name"], given_name)
                self.assertEqual(civil_by_id[national_id]["surname"], surname)
                self.assertEqual(civil_by_id[national_id]["birth_date"], birth_date)
                self.assertEqual(civil_by_id[national_id]["life_stage"], life_stage)
                self.assertEqual(civil_by_id[national_id]["deceased"], deceased)

    def test_guided_demo_personas_have_credible_source_rows(self) -> None:
        civil_by_id = {
            row[0]: {
                "given_name": row[1],
                "surname": row[2],
                "birth_date": row[3],
                "life_stage": row[4],
                "deceased": row[5],
                "district": row[6],
            }
            for row in self.generator.data_rows(self.generator.CIVIL_ROWS)
        }
        enrollment_by_id = {row[3]: row for row in self.generator.data_rows(self.generator.ENROLLMENTS)}
        health_by_id = {row["national_id"]: row for row in self.generator.HEALTH_ROWS}
        agri = load_module(AGRI_GENERATOR_PATH)
        farmers_by_id = {row["farmer_id"]: row for row in agri.FARMERS}

        self.assertEqual(
            civil_by_id["NID-1001"],
            {
                "given_name": "Miguel",
                "surname": "Santos",
                "birth_date": "2016-01-15",
                "life_stage": "child",
                "deceased": "false",
                "district": "north",
            },
        )
        self.assertEqual(civil_by_id["NID-2001"]["life_stage"], "adult")
        self.assertEqual(civil_by_id["NID-2001"]["deceased"], "false")
        self.assertEqual(enrollment_by_id["NID-1001"][4:6], ["CHILD_SUPPORT", "active"])
        self.assertTrue(self._health_available(health_by_id["NID-1001"]))
        self.assertFalse(self._health_available(health_by_id["NID-1002"]))
        self.assertEqual(farmers_by_id["FARMER-1001"]["given_name"], "Amina")
        self.assertEqual(farmers_by_id["FARMER-1001"]["family_name"], "Kone")
        self.assertEqual(farmers_by_id["FARMER-1002"]["given_name"], "Bako")
        self.assertEqual(farmers_by_id["FARMER-1003"]["given_name"], "Chipo")

    def test_public_demo_names_do_not_regress_to_misleading_civil_status(self) -> None:
        self.assertIn("life_stage", self.generator.CIVIL_ROWS[0])
        self.assertNotIn("civil_status", self.generator.CIVIL_ROWS[0])

        public_files = [
            self.generator.ROOT / "config" / "coolify" / "relay" / "civil-registry-relay.yaml",
            self.generator.ROOT / "config" / "coolify" / "relay" / "civil-registry-relay.metadata.yaml",
            self.generator.ROOT / "config" / "static-metadata" / "metadata.yaml",
        ]
        for path in public_files:
            with self.subTest(path=path.relative_to(self.generator.ROOT)):
                text = path.read_text(encoding="utf-8")
                self.assertNotRegex(text, r"\bname:\s*civil_status\b")
                self.assertNotRegex(text, r"\bcivil-status\b(?!-records)")

    def test_relay_metadata_does_not_publish_unbacked_notary_services(self) -> None:
        public_metadata_paths = [
            self.generator.ROOT / "config" / "static-metadata" / "metadata.yaml",
            *sorted((self.generator.ROOT / "config" / "coolify" / "relay").glob("*.metadata.yaml")),
        ]
        for path in public_metadata_paths:
            with self.subTest(path=path.relative_to(self.generator.ROOT)):
                text = path.read_text(encoding="utf-8")
                self.assertNotIn("evidence_offerings:", text)
                self.assertNotIn("kind: registry-notary", text)

    def test_v1_notary_outcomes_are_encoded_by_fixture_facts(self) -> None:
        civil_by_id = {row[0]: row for row in self.generator.data_rows(self.generator.CIVIL_ROWS)}
        enrollment_by_id = {row[3]: row for row in self.generator.data_rows(self.generator.ENROLLMENTS)}
        health_by_id = {row["national_id"]: row for row in self.generator.HEALTH_ROWS}

        for national_id, expected in self.expected_outcomes.items():
            with self.subTest(national_id=national_id):
                alive = civil_by_id[national_id][5] == "false"
                health = self._health_available(health_by_id[national_id])
                social_active = enrollment_by_id.get(national_id, [None] * 6)[5] == "active"
                self.assertEqual(alive, expected["alive"])
                self.assertEqual(health, expected["health"])
                self.assertEqual(alive and health and social_active, expected["combined"])

    def test_fixture_relationships_remain_valid(self) -> None:
        self.generator.validate_fixture_coverage()

    def test_civil_refresh_model_has_event_records_and_controls(self) -> None:
        person_by_nid = self._rows_by(self.generator.CIVIL_PERSON_DETAILS, "national_id")
        records_by_id = self._rows_by(self.generator.CIVIL_STATUS_RECORDS, "record_id")
        births_by_id = self._rows_by(self.generator.BIRTH_EVENTS, "event_id")
        deaths_by_id = self._rows_by(self.generator.DEATH_EVENTS, "event_id")
        marriages_by_id = self._rows_by(self.generator.MARRIAGE_EVENTS, "event_id")
        certificates_by_record = self._rows_by(self.generator.CERTIFICATES, "record_id")
        relationships_by_id = self._rows_by(self.generator.RELATIONSHIPS, "relationship_id")

        self.assertEqual(person_by_nid["NID-1001"]["sex"], "M")
        self.assertEqual(records_by_id["CSR-BIRTH-1001"]["event_id"], "BE-1001")
        self.assertEqual(births_by_id["BE-1001"]["child_person_id"], "CP-1001")
        self.assertEqual(births_by_id["BE-1001"]["father_person_id"], "CP-2002")
        self.assertEqual(certificates_by_record["CSR-BIRTH-1001"]["certificate_number"], "CERT-B-1001")
        self.assertEqual(relationships_by_id["REL-1001-MOTHER"]["related_person_id"], "CP-2001")
        self.assertEqual(relationships_by_id["REL-1001-FATHER"]["related_person_id"], "CP-2002")
        self.assertEqual(records_by_id["CSR-DEATH-1003"]["event_id"], "DE-1003")
        self.assertEqual(deaths_by_id["DE-1003"]["deceased_person_id"], "CP-1003")
        self.assertEqual(records_by_id["CSR-MARRIAGE-2001"]["event_id"], "ME-2001")
        self.assertEqual(marriages_by_id["ME-2001"]["spouse_1_person_id"], "CP-2001")
        self.assertEqual(certificates_by_record["CSR-MARRIAGE-2001"]["certificate_number"], "CERT-M-2001")

        ambiguous = [row for row in person_by_nid.values() if row["given_name"] == "Miguel" and row["surname"] == "Santos" and row["birth_date"] == "2016-01-15"]
        self.assertEqual({row["national_id"] for row in ambiguous}, {"NID-1001", "NID-1011"})
        self.assertEqual({row["place_of_birth"] for row in ambiguous}, {"North City", "South Town"})
        self.assertFalse(any(row["subject_person_id"] == "CP-1011" for row in relationships_by_id.values()))

    def test_social_refresh_model_splits_household_evidence_from_benefits(self) -> None:
        self.assertEqual(
            self.generator.HOUSEHOLDS[0],
            ["household_id", "national_id", "district", "poverty_score", "eligibility_band", "household_size", "active_members", "deceased_member_count", "observed_at"],
        )
        self.assertEqual(
            self.generator.ENROLLMENTS[0],
            ["enrollment_id", "household_id", "person_id", "national_id", "program_code", "status", "benefit_amount", "enrolled_on", "observed_at"],
        )

        memberships_by_person = self._rows_by(self.generator.GROUP_MEMBERSHIPS, "person_id")
        profiles_by_household = self._rows_by(self.generator.SOCIO_ECONOMIC_PROFILES, "household_id")
        scoring_by_profile = self._rows_by(self.generator.SCORING_EVENTS, "profile_id")
        programs_by_code = self._rows_by(self.generator.PROGRAMS, "program_code")
        entitlements_by_enrollment = self._rows_by(self.generator.ENTITLEMENTS, "enrollment_id")
        payments_by_entitlement = self._rows_by(self.generator.PAYMENT_EVENTS, "entitlement_id")
        functioning_by_nid = self._rows_by(self.generator.FUNCTIONING_PROFILES, "national_id")
        determinations_by_nid = self._rows_by(self.generator.DISABILITY_DETERMINATIONS, "national_id")

        self.assertEqual(memberships_by_person["PER-1001"]["relationship_type"], "child")
        self.assertEqual(profiles_by_household["HH-100"]["instrument"], "PMT-CHILD-2025")
        self.assertEqual(scoring_by_profile["SEP-100"]["score_band"], "priority")
        self.assertEqual(programs_by_code["CHILD_SUPPORT"]["display_name"], "Child Support Grant")
        self.assertEqual(entitlements_by_enrollment["ENR-100"]["entitlement_status"], "active")
        self.assertEqual(payments_by_entitlement["ENT-100"]["status"], "paid")
        self.assertEqual(functioning_by_nid["NID-1006"]["instrument_code"], "WG-SS-2025")
        self.assertIs(functioning_by_nid["NID-1006"]["disability_identifier_met"], True)
        self.assertEqual(functioning_by_nid["NID-1006"]["domains_triggering_identifier"], "mobility;self_care")
        self.assertEqual(determinations_by_nid["NID-1006"]["determination_status"], "approved")
        self.assertEqual(determinations_by_nid["NID-1006"]["support_category"], "top_up")

    def test_health_projection_is_reframed_for_applicant_key_compatibility(self) -> None:
        self.assertEqual(self.generator.HEALTH_PROJECTION_NAME, "ApplicantServiceAvailabilityProjection")
        self.assertIs(self.generator.HEALTH_ROWS, self.generator.APPLICANT_SERVICE_AVAILABILITY_PROJECTION)
        self.assertTrue(all("national_id" in row for row in self.generator.HEALTH_ROWS))
        self.assertTrue(all("facility_name" in row for row in self.generator.HEALTH_ROWS))
        self.assertTrue(all("observed_at" in row for row in self.generator.HEALTH_ROWS))

    def test_live_baseline_sources_carry_row_level_observed_at(self) -> None:
        self.assertEqual(self.generator.CIVIL_ROWS[0][-1], "observed_at")
        self.assertEqual(self.generator.HOUSEHOLDS[0][-1], "observed_at")
        self.assertEqual(self.generator.PERSONS[0][-1], "observed_at")
        self.assertEqual(self.generator.ENROLLMENTS[0][-1], "observed_at")
        self.assertEqual(self.generator.FUNCTIONING_PROFILES[0][-1], "observed_at")
        self.assertEqual(self.generator.DISABILITY_DETERMINATIONS[0][-1], "observed_at")

        sources = [
            (row[0], row[-1])
            for row in self.generator.data_rows(self.generator.CIVIL_ROWS)
        ] + [
            (row[1], row[-1])
            for row in self.generator.data_rows(self.generator.HOUSEHOLDS)
        ] + [
            (row[2], row[-1])
            for row in self.generator.data_rows(self.generator.PERSONS)
        ] + [
            (row[3], row[-1])
            for row in self.generator.data_rows(self.generator.ENROLLMENTS)
        ] + [
            (row[2], row[-1])
            for row in self.generator.data_rows(self.generator.FUNCTIONING_PROFILES)
        ] + [
            (row[2], row[-1])
            for row in self.generator.data_rows(self.generator.DISABILITY_DETERMINATIONS)
        ] + [
            (row["national_id"], row["observed_at"])
            for row in self.generator.HEALTH_ROWS
        ]
        self.assertTrue(sources)
        for national_id, value in sources:
            with self.subTest(national_id=national_id, value=value):
                self.assertEqual(value, self.generator.observed_at_for_national_id(national_id))
                if national_id == self.generator.MISSING_SOURCE_OBSERVED_AT_NATIONAL_ID:
                    self.assertEqual(value, "")
                    continue
                parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
                self.assertIsNotNone(parsed.tzinfo)
                if national_id == "NID-1010":
                    self.assertEqual(value, self.generator.STALE_SOURCE_OBSERVED_AT)

    def test_refresh_persona_invariants_cover_source_outcomes(self) -> None:
        personas = self.generator.FIXTURE_PERSONAS
        self.assertEqual(
            {persona["expected_outcome"] for persona in personas.values()},
            {"positive", "negative", "ambiguous_match", "stale", "expired", "policy_denied"},
        )

        civil_by_record = self._rows_by(self.generator.CIVIL_STATUS_RECORDS, "record_id")
        deaths_by_id = self._rows_by(self.generator.DEATH_EVENTS, "event_id")
        scoring_by_id = self._rows_by(self.generator.SCORING_EVENTS, "scoring_id")
        entitlements_by_id = self._rows_by(self.generator.ENTITLEMENTS, "entitlement_id")
        enrollments_by_id = self._rows_by(self.generator.ENROLLMENTS, "enrollment_id")
        relationships_by_id = self._rows_by(self.generator.RELATIONSHIPS, "relationship_id")
        details_by_nid = self._rows_by(self.generator.CIVIL_PERSON_DETAILS, "national_id")

        positive = personas["positive_child_benefit"]
        self.assertEqual(enrollments_by_id[positive["enrollment_id"]]["person_id"], positive["person_id"])
        self.assertEqual(civil_by_record[positive["civil_record_id"]]["registration_status"], "registered")
        self.assertEqual(relationships_by_id[positive["relationship_id"]]["relationship_status"], "established")
        self.assertEqual(scoring_by_id[positive["scoring_id"]]["scoring_status"], "current")
        self.assertEqual(entitlements_by_id[positive["entitlement_id"]]["entitlement_status"], "active")

        negative = personas["negative_deceased"]
        self.assertEqual(deaths_by_id[negative["death_event_id"]]["deceased_person_id"], "CP-1003")
        self.assertEqual(details_by_nid[negative["national_id"]]["deceased"], "true")

        ambiguous = personas["ambiguous_demographic"]
        self.assertNotEqual(
            details_by_nid[ambiguous["national_id"]]["place_of_birth"],
            details_by_nid[ambiguous["ambiguous_with"]]["place_of_birth"],
        )
        self.assertEqual(details_by_nid[ambiguous["national_id"]]["birth_date"], details_by_nid[ambiguous["ambiguous_with"]]["birth_date"])

        stale = personas["stale_welfare"]
        self.assertEqual(scoring_by_id[stale["scoring_id"]]["scoring_status"], "stale")
        self.assertLess(scoring_by_id[stale["scoring_id"]]["valid_until"], dt.date(2026, 1, 1))

        expired = personas["expired_entitlement"]
        self.assertEqual(enrollments_by_id[expired["enrollment_id"]]["person_id"], expired["person_id"])
        self.assertEqual(entitlements_by_id[expired["entitlement_id"]]["entitlement_status"], "expired")
        self.assertLess(entitlements_by_id[expired["entitlement_id"]]["coverage_end"], dt.date(2026, 1, 1))

        denied = personas["policy_denied"]
        self.assertEqual(enrollments_by_id[denied["enrollment_id"]]["person_id"], denied["person_id"])
        self.assertEqual(scoring_by_id[denied["scoring_id"]]["scoring_status"], "policy_denied")
        self.assertEqual(entitlements_by_id[denied["entitlement_id"]]["entitlement_status"], "policy_denied")

    def test_generated_outputs_match_fixture_source(self) -> None:
        civil_path = self.generator.DATA_DIR / "civil" / "civil-persons.csv"
        social_path = self.generator.DATA_DIR / "social-protection" / "social-protection.xlsx"
        health_path = self.generator.DATA_DIR / "health" / "health-facilities.parquet"

        with civil_path.open(newline="", encoding="utf-8") as handle:
            civil_rows = list(csv.reader(handle))
        self.assertEqual(
            self._normalize_observed_at_rows(civil_rows, national_id_column="national_id"),
            self._normalize_observed_at_rows(self.generator.CIVIL_ROWS, national_id_column="national_id"),
        )

        workbook = load_workbook(social_path, data_only=True)
        expected_sheets = {
            "Households": self.generator.HOUSEHOLDS,
            "Persons": self.generator.PERSONS,
            "Enrollments": self.generator.ENROLLMENTS,
            "DistrictGeometries": self.generator.DISTRICT_GEOMETRIES,
        }
        for sheet_name, expected_rows in expected_sheets.items():
            with self.subTest(sheet_name=sheet_name):
                sheet = workbook[sheet_name]
                actual_rows = [
                    [value.date() if isinstance(value, dt.datetime) else "" if value is None else value for value in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                if "observed_at" in expected_rows[0]:
                    self.assertEqual(
                        self._normalize_observed_at_rows(actual_rows, national_id_column="national_id"),
                        self._normalize_observed_at_rows(expected_rows, national_id_column="national_id"),
                    )
                else:
                    self.assertEqual(actual_rows, expected_rows)

        health_table = pq.read_table(health_path)
        health_rows = health_table.to_pylist()
        self.assertEqual(
            self._normalize_health_observed_at(health_rows),
            self._normalize_health_observed_at(self.generator.HEALTH_ROWS),
        )

    def _normalize_observed_at_rows(
        self,
        rows: list[list[object]],
        *,
        national_id_column: str,
    ) -> list[list[object]]:
        header = rows[0]
        if "observed_at" not in header:
            return rows
        observed_at_index = header.index("observed_at")
        national_id_index = header.index(national_id_column)
        normalized = [header]
        for row in rows[1:]:
            normalized_row = list(row)
            normalized_row[observed_at_index] = self._observed_at_category(str(row[national_id_index]))
            normalized.append(normalized_row)
        return normalized

    def _normalize_health_observed_at(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        normalized = []
        for row in rows:
            normalized_row = dict(row)
            normalized_row["observed_at"] = self._observed_at_category(str(row["national_id"]))
            normalized.append(normalized_row)
        return normalized

    def _observed_at_category(self, national_id: str) -> str:
        if national_id == "NID-1010":
            return "stale"
        if national_id == self.generator.MISSING_SOURCE_OBSERVED_AT_NATIONAL_ID:
            return "missing"
        return "fresh"

    def test_generator_writes_refreshed_fixture_model_outputs(self) -> None:
        original_data_dir = self.generator.DATA_DIR
        with tempfile.TemporaryDirectory() as tmp:
            self.generator.DATA_DIR = Path(tmp)
            try:
                self.generator.main()
                civil_dir = self.generator.DATA_DIR / "civil"
                for filename, expected_rows in [
                    ("civil-person-details.csv", self.generator.CIVIL_PERSON_DETAILS),
                    ("civil-identifiers.csv", self.generator.CIVIL_IDENTIFIERS),
                    ("birth-events.csv", self.generator.BIRTH_EVENTS),
                    ("death-events.csv", self.generator.DEATH_EVENTS),
                    ("marriage-events.csv", self.generator.MARRIAGE_EVENTS),
                    ("civil-status-records.csv", self.generator.CIVIL_STATUS_RECORDS),
                    ("certificates.csv", self.generator.CERTIFICATES),
                    ("relationships.csv", self.generator.RELATIONSHIPS),
                ]:
                    with self.subTest(filename=filename):
                        with (civil_dir / filename).open(newline="", encoding="utf-8") as handle:
                            self.assertEqual(list(csv.reader(handle)), self._stringified_rows(expected_rows))

                population_path = self.generator.DATA_DIR / "population" / "population-persons.csv"
                with population_path.open(newline="", encoding="utf-8") as handle:
                    self.assertEqual(
                        list(csv.reader(handle)),
                        self._stringified_rows(self.generator.POPULATION_PERSONS),
                    )

                workbook = load_workbook(self.generator.DATA_DIR / "social-protection" / "social-protection.xlsx", data_only=True)
                for sheet_name, expected_rows in {
                    "GroupMemberships": self.generator.GROUP_MEMBERSHIPS,
                    "SocioEconomicProfiles": self.generator.SOCIO_ECONOMIC_PROFILES,
                    "ScoringEvents": self.generator.SCORING_EVENTS,
                    "Programs": self.generator.PROGRAMS,
                    "Entitlements": self.generator.ENTITLEMENTS,
                    "PaymentEvents": self.generator.PAYMENT_EVENTS,
                    "FunctioningProfiles": self.generator.FUNCTIONING_PROFILES,
                    "DisabilityDeterminations": self.generator.DISABILITY_DETERMINATIONS,
                }.items():
                    with self.subTest(sheet_name=sheet_name):
                        actual_rows = [
                            [value.date() if isinstance(value, dt.datetime) else value for value in row]
                            for row in workbook[sheet_name].iter_rows(values_only=True)
                        ]
                        if "observed_at" in expected_rows[0]:
                            self.assertEqual(
                                self._normalize_observed_at_rows(actual_rows, national_id_column="national_id"),
                                self._normalize_observed_at_rows(expected_rows, national_id_column="national_id"),
                            )
                        else:
                            self.assertEqual(actual_rows, expected_rows)
            finally:
                self.generator.DATA_DIR = original_data_dir

    @staticmethod
    def _health_available(row: dict[str, object]) -> bool:
        return (
            row["license_status"] == "active"
            and row["pediatric_service_available"] is True
            and row["practitioner_credential_active"] is True
        )

    @staticmethod
    def _rows_by(rows: list[list[object]], key: str) -> dict[object, dict[str, object]]:
        header = rows[0]
        key_index = header.index(key)
        return {row[key_index]: dict(zip(header, row)) for row in rows[1:]}

    @staticmethod
    def _stringified_rows(rows: list[list[object]]) -> list[list[str]]:
        return [["" if value is None else str(value) for value in row] for row in rows]

if __name__ == "__main__":
    unittest.main()
